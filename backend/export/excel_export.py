"""
DC Feasibility Tool v4 - Excel Export
=====================================
Creates a scoped analyst workbook from the normalized report bundle.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from engine.models import ScenarioResult, Site
from export.report_data import build_report_context


def _header_style(cell, fill_color: str) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=fill_color.replace("#", "").upper())
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _body_style(cell) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (list, tuple, set)):
        parts = [str(_excel_value(item)) for item in value if _excel_value(item) not in ("", None)]
        return "; ".join(parts)
    return value


def _add_sheet(
    wb: Workbook,
    title: str,
    columns: list[tuple[str, str]],
    rows: list[dict[str, Any]],
    *,
    fill_color: str,
) -> None:
    ws = wb.create_sheet(title)
    ws.append([label for _, label in columns])
    for cell in ws[1]:
        _header_style(cell, fill_color)
    for row in rows:
        ws.append([_excel_value(row.get(key)) for key, _ in columns])
        for cell in ws[ws.max_row]:
            _body_style(cell)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _fit_columns(wb: Workbook) -> None:
    for sheet in wb.worksheets:
        for column_index, column in enumerate(sheet.iter_cols(), start=1):
            max_len = max(len(str(cell.value or "")) for cell in column)
            column_letter = get_column_letter(column_index)
            sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 60)


def _result_rows(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for result in results:
        scenario = result["scenario"]
        metrics = result["metrics"]
        status = result["status"]
        flags = result["feature_flags"]
        rows.append(
            {
                "site_id": result["site_id"],
                "site_name": result["site_name"],
                "global_rank": result["global_rank"],
                "rank_within_site": result["rank_within_site"],
                "selected_primary_rank": result["selected_primary_rank"],
                "is_primary": result["is_primary"],
                "is_displayed": result["is_displayed_in_current_output"],
                "result_key": result["result_key"],
                "label": result["label"],
                "load_type": scenario["load_type"],
                "cooling_type": scenario["cooling_type"],
                "redundancy": scenario["redundancy"],
                "density_scenario": scenario["density_scenario"],
                "backup_power": scenario["backup_power"],
                "preset_key": scenario["assumption_override_preset_key"],
                "preset_label": scenario["assumption_override_preset_label"],
                "pue_override": scenario["pue_override"],
                "compatible": result["compatible_combination"],
                "score": metrics["score"],
                "pue": metrics["pue"],
                "annual_pue": metrics["annual_pue"],
                "pue_source": metrics["pue_source"],
                "it_load_mw": metrics["it_load_mw"],
                "committed_it_mw": metrics["committed_it_mw"],
                "facility_power_mw": metrics["facility_power_mw"],
                "procurement_power_mw": metrics["procurement_power_mw"],
                "binding_constraint": metrics["binding_constraint"],
                "power_headroom_mw": metrics["power_headroom_mw"],
                "overtemperature_hours": metrics["overtemperature_hours"],
                "racks_deployed": metrics["racks_deployed"],
                "racks_by_power": metrics["racks_by_power"],
                "rack_density_kw": metrics["rack_density_kw"],
                "it_capacity_worst_mw": metrics["it_capacity_worst_mw"],
                "it_capacity_p99_mw": metrics["it_capacity_p99_mw"],
                "it_capacity_p90_mw": metrics["it_capacity_p90_mw"],
                "it_capacity_mean_mw": metrics["it_capacity_mean_mw"],
                "it_capacity_best_mw": metrics["it_capacity_best_mw"],
                "rag_status": status["rag_status"],
                "rag_reasons": _excel_value(status.get("rag_reasons") or []),
                "has_hourly_pue": flags["has_hourly_pue"],
                "has_it_capacity_spectrum": flags["has_it_capacity_spectrum"],
                "has_assumption_overrides": flags["has_assumption_overrides"],
                "override_count": len(result.get("applied_assumption_overrides") or []),
            }
        )
    return rows


def _selection_rows(studied_sites: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for site_bundle in studied_sites:
        summary = site_bundle["summary"]
        primary = site_bundle["results"]["primary"]
        rows.append(
            {
                "site_id": site_bundle["site_id"],
                "site_name": site_bundle["site_data"]["name"],
                "requested_primary_result_key": summary["requested_primary_result_key"],
                "resolved_primary_result_key": summary["resolved_primary_result_key"],
                "selected_result_label": primary["label"] if primary is not None else "",
                "selected_primary_rank": primary["selected_primary_rank"] if primary is not None else "",
                "available_result_count": summary["available_result_count"],
                "display_result_count": summary["display_result_count"],
                "alternative_count": summary["alternative_count"],
            }
        )
    return rows


def _site_rows(studied_sites: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for site_bundle in studied_sites:
        site_data = site_bundle["site_data"]
        location = site_data["location"]
        land = site_data["land"]
        power = site_data["power"]
        geometry = site_data["imported_geometry"]
        summary = site_bundle["summary"]
        primary = site_bundle["results"]["primary"]
        rows.append(
            {
                "site_id": site_bundle["site_id"],
                "site_name": site_data["name"],
                "site_type": site_data["site_type"],
                "country": location["country"],
                "city": location["city"],
                "coordinates": (
                    f'{location["latitude"]:.5f}, {location["longitude"]:.5f}'
                    if location["latitude"] is not None and location["longitude"] is not None
                    else ""
                ),
                "land_area_m2": land["land_area_m2"],
                "buildable_area_mode": land["buildable_area_mode"],
                "buildable_area_m2": land["buildable_area_m2"],
                "available_power_mw": power["available_power_mw"],
                "power_confirmed": power["power_confirmed"],
                "power_input_mode": power["power_input_mode"],
                "voltage": power["voltage"],
                "imported_geometry_present": geometry["present"],
                "imported_geometry_type": geometry["geometry_type"],
                "available_result_count": summary["available_result_count"],
                "display_result_count": summary["display_result_count"],
                "alternative_count": summary["alternative_count"],
                "requested_primary_result_key": summary["requested_primary_result_key"],
                "resolved_primary_result_key": summary["resolved_primary_result_key"],
                "selected_primary_label": primary["label"] if primary is not None else "",
                "avg_available_pue": summary["avg_available_pue"],
                "avg_available_score": summary["avg_available_score"],
                "max_available_it_load_mw": summary["max_available_it_load_mw"],
                "grid_context_status": site_bundle["grid_context"]["status"],
                "climate_status": site_bundle["climate"]["status"],
                "load_mix_status": site_bundle["load_mix"]["status"],
                "green_energy_status": site_bundle["green_energy"]["status"],
            }
        )
    return rows


def _metadata_rows(context: dict[str, Any]) -> list[dict[str, Any]]:
    report = context["report"]
    theme = context["theme"]
    scope = context["study_scope"]
    availability = context["report_bundle"]["analysis_availability"]
    policy = report["narrative_policy"]
    return [
        {"section": "Report", "key": "Report type", "value": report["type"]},
        {"section": "Report", "key": "Report title", "value": report["title"]},
        {"section": "Report", "key": "Layout mode label", "value": report["layout_mode_label"]},
        {"section": "Report", "key": "Generated at UTC", "value": report["generated_at_utc"]},
        {"section": "Narrative Policy", "key": "Mode", "value": policy["mode"]},
        {"section": "Narrative Policy", "key": "Max paragraphs", "value": policy["max_paragraphs"]},
        {"section": "Narrative Policy", "key": "Traceability", "value": policy["traceability"]},
        {"section": "Theme", "key": "Primary color", "value": theme["primary_color"]},
        {"section": "Theme", "key": "Secondary color", "value": theme["secondary_color"]},
        {"section": "Theme", "key": "Font family", "value": theme["font_family"]},
        {"section": "Theme", "key": "Logo URL", "value": theme["logo_url"]},
        {"section": "Study Scope", "key": "Cover scope label", "value": scope["cover_scope_label"]},
        {"section": "Study Scope", "key": "Studied site IDs", "value": _excel_value(scope["studied_site_ids"])},
        {"section": "Study Scope", "key": "Studied site names", "value": _excel_value(scope["studied_site_names"])},
        {"section": "Study Scope", "key": "Displayed result count", "value": scope["displayed_result_count"]},
        {"section": "Study Scope", "key": "Available result count", "value": scope["available_result_count"]},
        {"section": "Study Scope", "key": "Selected primary result count", "value": scope["selected_primary_result_count"]},
        {"section": "Analysis Availability", "key": "Grid context sites", "value": availability["grid_context_available_site_count"]},
        {"section": "Analysis Availability", "key": "Climate sites", "value": availability["climate_available_site_count"]},
        {"section": "Analysis Availability", "key": "Load mix sites", "value": availability["load_mix_available_site_count"]},
        {"section": "Analysis Availability", "key": "Green energy sites", "value": availability["green_energy_available_site_count"]},
        {"section": "Analysis Availability", "key": "Sites with alternatives", "value": availability["site_count_with_ranked_alternatives"]},
    ]


def _override_rows(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for result in results:
        for override in result.get("applied_assumption_overrides") or []:
            rows.append(
                {
                    "site_id": result["site_id"],
                    "site_name": result["site_name"],
                    "result_key": result["result_key"],
                    "result_label": result["label"],
                    "label": override.get("label"),
                    "scope_label": override.get("scope_label"),
                    "parameter_label": override.get("parameter_label"),
                    "unit": override.get("unit"),
                    "baseline_value": override.get("baseline_value"),
                    "effective_value": override.get("effective_value"),
                    "origin": override.get("origin"),
                    "source": override.get("source"),
                    "justification": override.get("justification"),
                    "updated_at_utc": override.get("updated_at_utc"),
                }
            )
    return rows


def _grid_rows(studied_sites: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    summary_rows: list[dict[str, Any]] = []
    asset_rows: list[dict[str, Any]] = []
    for site_bundle in studied_sites:
        block = site_bundle["grid_context"]
        selected = block.get("selected") or {}
        summary = selected.get("summary") or {}
        score = selected.get("score") or {}
        summary_rows.append(
            {
                "site_id": site_bundle["site_id"],
                "site_name": site_bundle["site_data"]["name"],
                "status": block["status"],
                "message": block.get("message"),
                "radius_km": summary.get("radius_km"),
                "asset_count": selected.get("asset_count"),
                "nearby_line_count": summary.get("nearby_line_count"),
                "nearby_substation_count": summary.get("nearby_substation_count"),
                "max_voltage_kv": summary.get("max_voltage_kv"),
                "overall_score": score.get("overall_score"),
                "confidence": selected.get("confidence"),
                "source_layers": _excel_value(selected.get("source_layers") or []),
            }
        )
        for asset in selected.get("assets") or []:
            asset_rows.append(
                {
                    "site_id": site_bundle["site_id"],
                    "site_name": site_bundle["site_data"]["name"],
                    "asset_id": asset.get("asset_id"),
                    "asset_type": asset.get("asset_type"),
                    "name": asset.get("name"),
                    "operator": asset.get("operator"),
                    "voltage_kv": asset.get("voltage_kv"),
                    "distance_km": asset.get("distance_km"),
                    "confidence": asset.get("confidence"),
                    "source": asset.get("source"),
                }
            )
    return summary_rows, asset_rows


def _climate_rows(studied_sites: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    summary_rows: list[dict[str, Any]] = []
    free_cooling_rows: list[dict[str, Any]] = []
    for site_bundle in studied_sites:
        block = site_bundle["climate"]
        weather = block.get("weather_status") or {}
        analysis = block.get("analysis") or {}
        stats = analysis.get("temperature_stats") or {}
        summary_rows.append(
            {
                "site_id": site_bundle["site_id"],
                "site_name": site_bundle["site_data"]["name"],
                "status": block["status"],
                "message": block.get("message"),
                "weather_source": weather.get("source"),
                "source_type": weather.get("source_type"),
                "hours": weather.get("hours"),
                "has_humidity": weather.get("has_humidity"),
                "years_averaged": _excel_value(weather.get("years_averaged") or []),
                "temperature_mean_c": stats.get("mean"),
                "temperature_min_c": stats.get("min"),
                "temperature_max_c": stats.get("max"),
                "temperature_p99_c": stats.get("p99"),
                "cooling_types_analyzed": _excel_value(analysis.get("cooling_types_analyzed") or []),
            }
        )
        for item in analysis.get("free_cooling") or []:
            free_cooling_rows.append(
                {
                    "site_id": site_bundle["site_id"],
                    "site_name": site_bundle["site_data"]["name"],
                    "cooling_type": item.get("cooling_type"),
                    "threshold_description": item.get("threshold_description"),
                    "free_cooling_hours": item.get("free_cooling_hours"),
                    "free_cooling_fraction": item.get("free_cooling_fraction"),
                    "partial_hours": item.get("partial_hours"),
                    "mechanical_hours": item.get("mechanical_hours"),
                    "suitability": item.get("suitability"),
                }
            )
    return summary_rows, free_cooling_rows


def _load_mix_rows(studied_sites: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    summary_rows: list[dict[str, Any]] = []
    candidate_rows: list[dict[str, Any]] = []
    for site_bundle in studied_sites:
        block = site_bundle["load_mix"]
        result = block.get("result") or {}
        candidates = result.get("top_candidates") or []
        top = candidates[0] if candidates else {}
        summary_rows.append(
            {
                "site_id": site_bundle["site_id"],
                "site_name": site_bundle["site_data"]["name"],
                "status": block["status"],
                "message": block.get("message"),
                "result_key": block.get("result_key"),
                "allowed_load_types": _excel_value(result.get("allowed_load_types") or []),
                "total_it_mw": result.get("total_it_mw"),
                "cooling_type": result.get("cooling_type"),
                "density_scenario": result.get("density_scenario"),
                "step_pct": result.get("step_pct"),
                "total_candidates_evaluated": result.get("total_candidates_evaluated"),
                "top_candidate_score": top.get("score"),
                "top_candidate_blended_pue": top.get("blended_pue"),
                "top_candidate_total_racks": top.get("total_racks"),
            }
        )
        for candidate in candidates:
            candidate_rows.append(
                {
                    "site_id": site_bundle["site_id"],
                    "site_name": site_bundle["site_data"]["name"],
                    "candidate_rank": candidate.get("rank"),
                    "score": candidate.get("score"),
                    "blended_pue": candidate.get("blended_pue"),
                    "total_racks": candidate.get("total_racks"),
                    "all_compatible": candidate.get("all_compatible"),
                    "trade_off_notes": _excel_value(candidate.get("trade_off_notes") or []),
                }
            )
    return summary_rows, candidate_rows


def _green_rows(studied_sites: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for site_bundle in studied_sites:
        block = site_bundle["green_energy"]
        result = block.get("result") or {}
        pvgis = block.get("pvgis_profile") or {}
        rows.append(
            {
                "site_id": site_bundle["site_id"],
                "site_name": site_bundle["site_data"]["name"],
                "status": block["status"],
                "message": block.get("message"),
                "result_key": block.get("result_key"),
                "renewable_fraction": result.get("renewable_fraction"),
                "overhead_coverage_fraction": result.get("overhead_coverage_fraction"),
                "co2_avoided_tonnes": result.get("co2_avoided_tonnes"),
                "total_facility_kwh": result.get("total_facility_kwh"),
                "total_it_kwh": result.get("total_it_kwh"),
                "total_overhead_kwh": result.get("total_overhead_kwh"),
                "total_pv_generation_kwh": result.get("total_pv_generation_kwh"),
                "total_grid_import_kwh": result.get("total_grid_import_kwh"),
                "pv_capacity_kwp": result.get("pv_capacity_kwp"),
                "bess_capacity_kwh": result.get("bess_capacity_kwh"),
                "fuel_cell_capacity_kw": result.get("fuel_cell_capacity_kw"),
                "annual_pue": result.get("annual_pue"),
                "pue_source": result.get("pue_source"),
                "committed_it_mw": result.get("committed_it_mw"),
                "pv_profile_source": result.get("pv_profile_source"),
                "pvgis_profile_key": result.get("pvgis_profile_key"),
                "pv_profile_name": block.get("pv_profile_name"),
                "pvgis_years_averaged": _excel_value(pvgis.get("years_averaged") or []),
            }
        )
    return rows


def _appendix_rows(
    studied_sites: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    summary_rows: list[dict[str, Any]] = []
    note_rows: list[dict[str, Any]] = []
    table_rows: list[dict[str, Any]] = []
    for site_bundle in studied_sites:
        primary = site_bundle["results"]["primary"]
        if primary is None:
            continue
        for block in site_bundle["chapters"]["deep_dive"].get("advanced_blocks") or []:
            for index, item in enumerate(block.get("summary_items") or [], start=1):
                summary_rows.append(
                    {
                        "site_id": site_bundle["site_id"],
                        "site_name": site_bundle["site_data"]["name"],
                        "result_key": primary["result_key"],
                        "result_label": primary["label"],
                        "block_key": block["key"],
                        "block_title": block["title"],
                        "item_order": index,
                        "item_label": item.get("label"),
                        "item_value": item.get("value"),
                    }
                )
            for index, note in enumerate(block.get("notes") or [], start=1):
                note_rows.append(
                    {
                        "site_id": site_bundle["site_id"],
                        "site_name": site_bundle["site_data"]["name"],
                        "result_key": primary["result_key"],
                        "result_label": primary["label"],
                        "block_key": block["key"],
                        "block_title": block["title"],
                        "note_order": index,
                        "note": note,
                    }
                )
            for table in block.get("tables") or []:
                labels = {column["key"]: column["label"] for column in table.get("columns") or []}
                for row_index, row in enumerate(table.get("rows") or [], start=1):
                    for column_key, value in row.items():
                        table_rows.append(
                            {
                                "site_id": site_bundle["site_id"],
                                "site_name": site_bundle["site_data"]["name"],
                                "result_key": primary["result_key"],
                                "result_label": primary["label"],
                                "block_key": block["key"],
                                "block_title": block["title"],
                                "table_title": table["title"],
                                "row_index": row_index,
                                "column_key": column_key,
                                "column_label": labels.get(column_key, column_key),
                                "value": value,
                            }
                        )
    return summary_rows, note_rows, table_rows


def build_excel_bytes(
    report_type: str,
    primary_color: str,
    secondary_color: str,
    font_family: str,
    logo_url: str | None,
    site_entries: list[tuple[str, Site]],
    scenario_results: list[ScenarioResult],
    layout_mode: str = "presentation_16_9",
    studied_site_ids: list[str] | None = None,
    primary_result_keys: dict[str, str] | None = None,
    load_mix_results: dict[str, object] | None = None,
    green_energy_results: dict[str, object] | None = None,
    include_all_scenarios: bool = True,
) -> bytes:
    context = build_report_context(
        report_type=report_type,
        primary_color=primary_color,
        secondary_color=secondary_color,
        font_family=font_family,
        logo_url=logo_url,
        site_entries=site_entries,
        scenario_results=scenario_results,
        layout_mode=layout_mode,
        studied_site_ids=studied_site_ids,
        primary_result_keys=primary_result_keys,
        load_mix_results=load_mix_results,
        green_energy_results=green_energy_results,
        include_all_scenarios=include_all_scenarios,
    )
    bundle = context["report_bundle"]
    studied_sites = bundle["studied_sites"]
    primary_results = bundle["selected_primary_results"]
    all_results = bundle["all_ranked_results"]
    availability = bundle["analysis_availability"]

    wb = Workbook()
    ws = wb.active
    if not isinstance(ws, Worksheet):
        ws = wb.create_sheet("Summary", 0)
    else:
        ws.title = "Summary"
    ws.append(["Metric", "Value"])
    for cell in ws[1]:
        _header_style(cell, primary_color)

    top_primary = primary_results[0] if primary_results else None
    top_available = all_results[0] if all_results else None
    summary_rows = [
        ("Report Type", context["report"]["title"]),
        ("Layout Mode", context["report"]["layout_mode_label"]),
        ("Generated At UTC", context["report"]["generated_at_utc"]),
        ("Sites Included", context["summary"]["site_count"]),
        ("Scenario Results Included", context["summary"]["scenario_count"]),
        ("Full Scenario Matrix Rows", context["summary"]["available_scenario_count"]),
        ("Primary Results Included", context["summary"]["primary_result_count"]),
        ("Average PUE", context["summary"]["avg_pue"]),
        ("Full-Matrix Average PUE", context["summary"]["available_avg_pue"]),
        ("Maximum IT Load (MW)", context["summary"]["max_it_load_mw"]),
        ("Full-Matrix Max IT Load (MW)", context["summary"]["available_max_it_load_mw"]),
        ("Top Primary Scenario", top_primary["label"] if top_primary is not None else ""),
        ("Top Available Scenario", top_available["label"] if top_available is not None else ""),
        ("Grid Context Sites Available", availability["grid_context_available_site_count"]),
        ("Climate Sites Available", availability["climate_available_site_count"]),
        ("Load Mix Sites Available", availability["load_mix_available_site_count"]),
        ("Green Energy Sites Available", availability["green_energy_available_site_count"]),
    ]
    for metric, value in summary_rows:
        ws.append([metric, _excel_value(value)])
        for cell in ws[ws.max_row]:
            _body_style(cell)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    _add_sheet(
        wb,
        "Metadata",
        [("section", "Section"), ("key", "Key"), ("value", "Value")],
        _metadata_rows(context),
        fill_color=secondary_color,
    )
    _add_sheet(
        wb,
        "Selections",
        [
            ("site_id", "Site ID"),
            ("site_name", "Site Name"),
            ("requested_primary_result_key", "Requested Primary Result Key"),
            ("resolved_primary_result_key", "Resolved Primary Result Key"),
            ("selected_result_label", "Selected Result Label"),
            ("selected_primary_rank", "Selected Primary Rank"),
            ("available_result_count", "Available Result Count"),
            ("display_result_count", "Displayed Result Count"),
            ("alternative_count", "Alternative Count"),
        ],
        _selection_rows(studied_sites),
        fill_color=secondary_color,
    )
    _add_sheet(
        wb,
        "Sites",
        [
            ("site_id", "Site ID"),
            ("site_name", "Site Name"),
            ("site_type", "Site Type"),
            ("country", "Country"),
            ("city", "City"),
            ("coordinates", "Coordinates"),
            ("land_area_m2", "Land Area m2"),
            ("buildable_area_mode", "Buildable Area Mode"),
            ("buildable_area_m2", "Buildable Area m2"),
            ("available_power_mw", "Available Power MW"),
            ("power_confirmed", "Power Confirmed"),
            ("power_input_mode", "Power Input Mode"),
            ("voltage", "Voltage"),
            ("imported_geometry_present", "Imported Geometry Present"),
            ("imported_geometry_type", "Imported Geometry Type"),
            ("available_result_count", "Available Result Count"),
            ("display_result_count", "Displayed Result Count"),
            ("alternative_count", "Alternative Count"),
            ("requested_primary_result_key", "Requested Primary Result Key"),
            ("resolved_primary_result_key", "Resolved Primary Result Key"),
            ("selected_primary_label", "Selected Primary Label"),
            ("avg_available_pue", "Average Available PUE"),
            ("avg_available_score", "Average Available Score"),
            ("max_available_it_load_mw", "Max Available IT Load MW"),
            ("grid_context_status", "Grid Context Status"),
            ("climate_status", "Climate Status"),
            ("load_mix_status", "Load Mix Status"),
            ("green_energy_status", "Green Energy Status"),
        ],
        _site_rows(studied_sites),
        fill_color=secondary_color,
    )
    result_columns = [
        ("site_id", "Site ID"),
        ("site_name", "Site Name"),
        ("global_rank", "Global Rank"),
        ("rank_within_site", "Rank Within Site"),
        ("selected_primary_rank", "Selected Primary Rank"),
        ("is_primary", "Is Primary"),
        ("is_displayed", "Displayed In Main Report"),
        ("result_key", "Result Key"),
        ("label", "Result Label"),
        ("load_type", "Load Type"),
        ("cooling_type", "Cooling Type"),
        ("redundancy", "Redundancy"),
        ("density_scenario", "Density Scenario"),
        ("backup_power", "Backup Power"),
        ("preset_key", "Preset Key"),
        ("preset_label", "Preset Label"),
        ("pue_override", "Manual PUE Override"),
        ("compatible", "Compatible"),
        ("score", "Score"),
        ("pue", "PUE"),
        ("annual_pue", "Annual PUE"),
        ("pue_source", "PUE Source"),
        ("it_load_mw", "IT Load MW"),
        ("committed_it_mw", "Committed IT MW"),
        ("facility_power_mw", "Facility Power MW"),
        ("procurement_power_mw", "Procurement Power MW"),
        ("binding_constraint", "Binding Constraint"),
        ("power_headroom_mw", "Power Headroom MW"),
        ("overtemperature_hours", "Overtemperature Hours"),
        ("racks_deployed", "Racks Deployed"),
        ("racks_by_power", "Racks by Power"),
        ("rack_density_kw", "Rack Density kW"),
        ("it_capacity_worst_mw", "IT Capacity Worst MW"),
        ("it_capacity_p99_mw", "IT Capacity P99 MW"),
        ("it_capacity_p90_mw", "IT Capacity P90 MW"),
        ("it_capacity_mean_mw", "IT Capacity Mean MW"),
        ("it_capacity_best_mw", "IT Capacity Best MW"),
        ("rag_status", "RAG Status"),
        ("rag_reasons", "RAG Reasons"),
        ("has_hourly_pue", "Has Hourly PUE"),
        ("has_it_capacity_spectrum", "Has IT Capacity Spectrum"),
        ("has_assumption_overrides", "Has Assumption Overrides"),
        ("override_count", "Override Count"),
    ]
    _add_sheet(wb, "Primary Results", result_columns, _result_rows(primary_results), fill_color=primary_color)
    _add_sheet(wb, "Scenarios", result_columns, _result_rows(all_results), fill_color=primary_color)

    overrides = _override_rows(all_results)
    if overrides:
        _add_sheet(
            wb,
            "Scenario Overrides",
            [
                ("site_id", "Site ID"),
                ("site_name", "Site Name"),
                ("result_key", "Result Key"),
                ("result_label", "Result Label"),
                ("label", "Override Label"),
                ("scope_label", "Scope"),
                ("parameter_label", "Parameter"),
                ("unit", "Unit"),
                ("baseline_value", "Baseline Value"),
                ("effective_value", "Effective Value"),
                ("origin", "Origin"),
                ("source", "Source"),
                ("justification", "Justification"),
                ("updated_at_utc", "Updated At UTC"),
            ],
            overrides,
            fill_color=secondary_color,
        )

    if availability["grid_context_available_site_count"] > 0:
        grid_summary_rows, grid_asset_rows = _grid_rows(studied_sites)
        _add_sheet(
            wb,
            "Grid Summary",
            [
                ("site_id", "Site ID"),
                ("site_name", "Site Name"),
                ("status", "Status"),
                ("message", "Message"),
                ("radius_km", "Radius km"),
                ("asset_count", "Asset Count"),
                ("nearby_line_count", "Nearby Line Count"),
                ("nearby_substation_count", "Nearby Substation Count"),
                ("max_voltage_kv", "Max Voltage kV"),
                ("overall_score", "Overall Score"),
                ("confidence", "Confidence"),
                ("source_layers", "Source Layers"),
            ],
            grid_summary_rows,
            fill_color=secondary_color,
        )
        if grid_asset_rows:
            _add_sheet(
                wb,
                "Grid Assets",
                [
                    ("site_id", "Site ID"),
                    ("site_name", "Site Name"),
                    ("asset_id", "Asset ID"),
                    ("asset_type", "Asset Type"),
                    ("name", "Asset Name"),
                    ("operator", "Operator"),
                    ("voltage_kv", "Voltage kV"),
                    ("distance_km", "Distance km"),
                    ("confidence", "Confidence"),
                    ("source", "Source"),
                ],
                grid_asset_rows,
                fill_color=secondary_color,
            )
 
    if availability["climate_available_site_count"] > 0:
        climate_summary_rows, climate_free_cooling_rows = _climate_rows(studied_sites)
        _add_sheet(
            wb,
            "Climate Summary",
            [
                ("site_id", "Site ID"),
                ("site_name", "Site Name"),
                ("status", "Status"),
                ("message", "Message"),
                ("weather_source", "Weather Source"),
                ("source_type", "Source Type"),
                ("hours", "Hours"),
                ("has_humidity", "Has Humidity"),
                ("years_averaged", "Years Averaged"),
                ("temperature_mean_c", "Mean C"),
                ("temperature_min_c", "Min C"),
                ("temperature_max_c", "Max C"),
                ("temperature_p99_c", "P99 C"),
                ("cooling_types_analyzed", "Cooling Types Analyzed"),
            ],
            climate_summary_rows,
            fill_color=secondary_color,
        )
        if climate_free_cooling_rows:
            _add_sheet(
                wb,
                "Climate Free Cooling",
                [
                    ("site_id", "Site ID"),
                    ("site_name", "Site Name"),
                    ("cooling_type", "Cooling Type"),
                    ("threshold_description", "Threshold"),
                    ("free_cooling_hours", "Free Cooling Hours"),
                    ("free_cooling_fraction", "Free Cooling Fraction"),
                    ("partial_hours", "Partial Hours"),
                    ("mechanical_hours", "Mechanical Hours"),
                    ("suitability", "Suitability"),
                ],
                climate_free_cooling_rows,
                fill_color=secondary_color,
            )

    if availability["load_mix_available_site_count"] > 0:
        load_mix_summary_rows, load_mix_candidate_rows = _load_mix_rows(studied_sites)
        _add_sheet(
            wb,
            "Load Mix",
            [
                ("site_id", "Site ID"),
                ("site_name", "Site Name"),
                ("status", "Status"),
                ("message", "Message"),
                ("result_key", "Result Key"),
                ("allowed_load_types", "Allowed Load Types"),
                ("total_it_mw", "Total IT MW"),
                ("cooling_type", "Cooling Type"),
                ("density_scenario", "Density Scenario"),
                ("step_pct", "Step Percent"),
                ("total_candidates_evaluated", "Candidates Evaluated"),
                ("top_candidate_score", "Top Candidate Score"),
                ("top_candidate_blended_pue", "Top Candidate Blended PUE"),
                ("top_candidate_total_racks", "Top Candidate Total Racks"),
            ],
            load_mix_summary_rows,
            fill_color=secondary_color,
        )
        if load_mix_candidate_rows:
            _add_sheet(
                wb,
                "Load Mix Candidates",
                [
                    ("site_id", "Site ID"),
                    ("site_name", "Site Name"),
                    ("candidate_rank", "Candidate Rank"),
                    ("score", "Score"),
                    ("blended_pue", "Blended PUE"),
                    ("total_racks", "Total Racks"),
                    ("all_compatible", "All Compatible"),
                    ("trade_off_notes", "Trade-Off Notes"),
                ],
                load_mix_candidate_rows,
                fill_color=secondary_color,
            )

    if availability["green_energy_available_site_count"] > 0:
        _add_sheet(
            wb,
            "Green Energy",
            [
                ("site_id", "Site ID"),
                ("site_name", "Site Name"),
                ("status", "Status"),
                ("message", "Message"),
                ("result_key", "Result Key"),
                ("renewable_fraction", "Renewable Fraction"),
                ("overhead_coverage_fraction", "Overhead Coverage Fraction"),
                ("co2_avoided_tonnes", "CO2 Avoided Tonnes"),
                ("total_facility_kwh", "Total Facility kWh"),
                ("total_it_kwh", "Total IT kWh"),
                ("total_overhead_kwh", "Total Overhead kWh"),
                ("total_pv_generation_kwh", "Total PV Generation kWh"),
                ("total_grid_import_kwh", "Total Grid Import kWh"),
                ("pv_capacity_kwp", "PV Capacity kWp"),
                ("bess_capacity_kwh", "BESS Capacity kWh"),
                ("fuel_cell_capacity_kw", "Fuel Cell Capacity kW"),
                ("annual_pue", "Annual PUE"),
                ("pue_source", "PUE Source"),
                ("committed_it_mw", "Committed IT MW"),
                ("pv_profile_source", "PV Profile Source"),
                ("pvgis_profile_key", "PVGIS Profile Key"),
                ("pv_profile_name", "PV Profile Name"),
                ("pvgis_years_averaged", "PVGIS Years Averaged"),
            ],
            _green_rows(studied_sites),
            fill_color=secondary_color,
        )

    appendix_summary_rows, appendix_note_rows, appendix_table_rows = _appendix_rows(studied_sites)
    if appendix_summary_rows:
        _add_sheet(
            wb,
            "Appx Summary",
            [
                ("site_id", "Site ID"),
                ("site_name", "Site Name"),
                ("result_key", "Result Key"),
                ("result_label", "Result Label"),
                ("block_key", "Block Key"),
                ("block_title", "Block Title"),
                ("item_order", "Item Order"),
                ("item_label", "Item Label"),
                ("item_value", "Item Value"),
            ],
            appendix_summary_rows,
            fill_color=primary_color,
        )
    if appendix_note_rows:
        _add_sheet(
            wb,
            "Appx Notes",
            [
                ("site_id", "Site ID"),
                ("site_name", "Site Name"),
                ("result_key", "Result Key"),
                ("result_label", "Result Label"),
                ("block_key", "Block Key"),
                ("block_title", "Block Title"),
                ("note_order", "Note Order"),
                ("note", "Note"),
            ],
            appendix_note_rows,
            fill_color=primary_color,
        )
    if appendix_table_rows:
        _add_sheet(
            wb,
            "Appx Tables",
            [
                ("site_id", "Site ID"),
                ("site_name", "Site Name"),
                ("result_key", "Result Key"),
                ("result_label", "Result Label"),
                ("block_key", "Block Key"),
                ("block_title", "Block Title"),
                ("table_title", "Table Title"),
                ("row_index", "Row Index"),
                ("column_key", "Column Key"),
                ("column_label", "Column Label"),
                ("value", "Value"),
            ],
            appendix_table_rows,
            fill_color=primary_color,
        )

    _fit_columns(wb)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
