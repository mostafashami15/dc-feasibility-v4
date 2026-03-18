import asyncio
import json
import shutil
import uuid
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from api import routes_export
from engine.models import (
    CoolingType,
    DensityScenario,
    ImportedGeometry,
    LoadType,
    PowerInputMode,
    PowerResult,
    RAGStatus,
    RedundancyLevel,
    Scenario,
    ScenarioResult,
    Site,
    SpaceResult,
)
from export import report_data
from export.excel_export import build_excel_bytes
from export.html_report import render_report_html
from export.report_data import (
    build_report_bundle,
    build_report_context,
    get_result_selection_key,
)

DEFAULT_EXPORT_STYLE = {
    "primary_color": "#1a365d",
    "secondary_color": "#2b6cb0",
    "font_family": "Inter, sans-serif",
    "logo_url": None,
}

_LOCAL_TEMP_DIRS: list[Path] = []
_LOCAL_TEMP_ROOT = Path("data") / ".pytest" / "export-temp"


@pytest.fixture(scope="module", autouse=True)
def cleanup_local_temp_dirs():
    yield
    for path in _LOCAL_TEMP_DIRS:
        shutil.rmtree(path, ignore_errors=True)
    _LOCAL_TEMP_DIRS.clear()
    shutil.rmtree(_LOCAL_TEMP_ROOT, ignore_errors=True)


@dataclass(frozen=True)
class ExportTestCase:
    report_type: str
    site_entries: list[tuple[str, Site]]
    scenario_results: list[ScenarioResult]
    studied_site_ids: list[str]
    primary_result_keys: dict[str, str]
    layout_mode: str = "presentation_16_9"
    grid_context_site_ids: set[str] = field(default_factory=set)
    weather_by_site: dict[str, dict[str, Any] | None] = field(default_factory=dict)
    load_mix_results: dict[str, dict[str, Any]] | None = None
    green_energy_results: dict[str, dict[str, Any]] | None = None

    def export_kwargs(self) -> dict[str, Any]:
        return {
            "report_type": self.report_type,
            **DEFAULT_EXPORT_STYLE,
            "site_entries": self.site_entries,
            "scenario_results": self.scenario_results,
            "layout_mode": self.layout_mode,
            "studied_site_ids": self.studied_site_ids,
            "primary_result_keys": self.primary_result_keys,
            "load_mix_results": self.load_mix_results,
            "green_energy_results": self.green_energy_results,
        }

    def install_optional_inputs(
        self,
        monkeypatch: pytest.MonkeyPatch,
        temp_dir: Path | None = None,
    ) -> None:
        grid_dir = temp_dir or make_local_temp_dir()
        grid_dir.mkdir(parents=True, exist_ok=True)
        for site_id, site in self.site_entries:
            if site_id in self.grid_context_site_ids:
                write_grid_context_cache(grid_dir, site_id, site.name)
        monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", grid_dir)
        weather_by_site = dict(self.weather_by_site)
        monkeypatch.setattr(
            report_data,
            "get_weather",
            lambda site_id: weather_by_site.get(site_id),
        )


def make_site(
    name: str,
    *,
    land_area_m2: float = 10000,
    available_power_mw: float = 20.0,
    power_confirmed: bool = False,
    latitude: float | None = None,
    longitude: float | None = None,
    imported_geometry: ImportedGeometry | None = None,
) -> Site:
    return Site(
        name=name,
        land_area_m2=land_area_m2,
        available_power_mw=available_power_mw,
        power_confirmed=power_confirmed,
        latitude=latitude,
        longitude=longitude,
        imported_geometry=imported_geometry,
        city="Milan",
        country="Italy",
    )


def make_result(
    site_id: str,
    site_name: str,
    *,
    load_type: LoadType = LoadType.COLOCATION_STANDARD,
    cooling_type: CoolingType = CoolingType.AIR_CHILLER_ECON,
    redundancy: RedundancyLevel = RedundancyLevel.TWO_N,
    density_scenario: DensityScenario = DensityScenario.TYPICAL,
    score: float = 80.0,
    it_load_mw: float = 12.0,
    pue_used: float = 1.32,
    annual_pue: float | None = None,
    overtemperature_hours: int | None = None,
    it_capacity_worst_mw: float | None = None,
    it_capacity_p99_mw: float | None = None,
    it_capacity_p90_mw: float | None = None,
    it_capacity_mean_mw: float | None = None,
    it_capacity_best_mw: float | None = None,
) -> ScenarioResult:
    return ScenarioResult(
        site_id=site_id,
        site_name=site_name,
        scenario=Scenario(
            load_type=load_type,
            cooling_type=cooling_type,
            redundancy=redundancy,
            density_scenario=density_scenario,
        ),
        space=SpaceResult(
            buildable_footprint_m2=5000.0,
            gross_building_area_m2=5000.0,
            it_whitespace_m2=2000.0,
            support_area_m2=3000.0,
            max_racks_by_space=1600,
            effective_racks=1400,
            whitespace_adjustment_factor=0.92,
            site_coverage_used=0.5,
            whitespace_ratio_used=0.4,
            rack_footprint_used=3.0,
            active_floors=1,
            floor_to_floor_height_used=4.5,
        ),
        power=PowerResult(
            it_load_mw=it_load_mw,
            facility_power_mw=round(it_load_mw * pue_used, 3),
            procurement_power_mw=round(it_load_mw * pue_used * 1.2, 3),
            racks_by_power=1500,
            racks_deployed=1200,
            rack_density_kw=10.0,
            binding_constraint="POWER",
            power_headroom_mw=2.5,
            eta_chain=0.92,
            pue_used=pue_used,
            procurement_factor=1.2,
            power_input_mode=PowerInputMode.OPERATIONAL,
            rag_status=RAGStatus.GREEN,
            rag_reasons=[],
        ),
        score=score,
        annual_pue=annual_pue,
        pue_source="hourly" if annual_pue is not None else "static",
        overtemperature_hours=overtemperature_hours,
        it_capacity_worst_mw=it_capacity_worst_mw,
        it_capacity_p99_mw=it_capacity_p99_mw,
        it_capacity_p90_mw=it_capacity_p90_mw,
        it_capacity_mean_mw=it_capacity_mean_mw,
        it_capacity_best_mw=it_capacity_best_mw,
    )


def write_grid_context_cache(tmp_path, site_id: str, site_name: str) -> None:
    cache_dir = tmp_path / site_id
    cache_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "site_id": site_id,
        "site_name": site_name,
        "latitude": 45.0,
        "longitude": 9.0,
        "analysis_grade": "screening_grade",
        "summary": {
            "radius_km": 5.0,
            "nearby_line_count": 1,
            "nearby_substation_count": 1,
            "nearest_line_km": 0.8,
            "nearest_substation_km": 1.2,
            "max_voltage_kv": 220.0,
            "high_voltage_assets_within_radius": 2,
        },
        "score": {
            "overall_score": 82.0,
            "voltage_score": 80.0,
            "distance_score": 75.0,
            "substation_score": 85.0,
            "evidence_score": 70.0,
            "notes": ["Fixture score"],
        },
        "assets": [
            {
                "asset_id": "line-1",
                "asset_type": "line",
                "name": "HV Line",
                "operator": "Fixture Grid",
                "voltage_kv": 220.0,
                "circuits": 2,
                "distance_km": 0.8,
                "geometry_type": "line",
                "coordinates": [[45.0, 9.0], [45.01, 9.02]],
                "source": "fixture",
                "confidence": "mapped_public",
            }
        ],
        "evidence_notes": [],
        "official_evidence": None,
        "official_context_notes": ["Fixture grid context"],
        "source_layers": ["fixture"],
        "confidence": "mapped_public",
        "generated_at_utc": "2026-03-13T10:00:00Z",
    }
    (cache_dir / "5km_v3.json").write_text(json.dumps(payload), encoding="utf-8")


def make_local_temp_dir() -> Path:
    _LOCAL_TEMP_ROOT.mkdir(parents=True, exist_ok=True)
    path = _LOCAL_TEMP_ROOT / f"tmp-{uuid.uuid4().hex}"
    path.mkdir(parents=True, exist_ok=True)
    _LOCAL_TEMP_DIRS.append(path)
    return path


def make_full_year_temperatures() -> tuple[list[float], list[float]]:
    month_hours = [744, 672, 744, 720, 744, 720, 744, 744, 720, 744, 720, 744]
    monthly_means = [4.0, 6.0, 9.0, 13.0, 18.0, 23.0, 27.0, 26.0, 21.0, 15.0, 9.0, 5.0]
    temperatures: list[float] = []
    humidities: list[float] = []
    for month_index, hours in enumerate(month_hours):
        mean_temp = monthly_means[month_index]
        for hour in range(hours):
            temperatures.append(mean_temp + ((hour % 24) - 12) * 0.12)
            humidities.append(55.0 - month_index)
    return temperatures, humidities


def make_weather_export_payload(
    *,
    temperatures: list[float],
    humidities: list[float] | None = None,
    years_averaged: list[int] | None = None,
    include_coordinates: bool = False,
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "temperatures": temperatures,
        "source": "Fixture Weather",
        "source_type": "manual_upload",
        "hours": len(temperatures),
        "years_averaged": years_averaged or [],
        "original_filename": "fixture-weather.csv",
    }
    if humidities is not None:
        payload["humidities"] = humidities
    if include_coordinates:
        payload["latitude"] = 45.0
        payload["longitude"] = 9.0
    return payload


def make_load_mix_export_payload(*, result_key: str | None = None) -> dict:
    return {
        "result_key": result_key,
        "result": {
            "total_it_mw": 12.0,
            "allowed_load_types": [
                LoadType.HYPERSCALE.value,
                LoadType.HPC.value,
                LoadType.AI_GPU.value,
            ],
            "cooling_type": CoolingType.WATER_CHILLER_ECON.value,
            "density_scenario": DensityScenario.TYPICAL.value,
            "step_pct": 10,
            "min_racks": 10,
            "total_candidates_evaluated": 6,
            "top_candidates": [
                {
                    "rank": 1,
                    "allocations": [
                        {
                            "load_type": LoadType.HYPERSCALE.value,
                            "share_pct": 50.0,
                            "it_load_mw": 6.0,
                            "rack_count": 600,
                            "rack_density_kw": 10.0,
                        },
                        {
                            "load_type": LoadType.HPC.value,
                            "share_pct": 30.0,
                            "it_load_mw": 3.6,
                            "rack_count": 360,
                            "rack_density_kw": 10.0,
                        },
                        {
                            "load_type": LoadType.AI_GPU.value,
                            "share_pct": 20.0,
                            "it_load_mw": 2.4,
                            "rack_count": 160,
                            "rack_density_kw": 15.0,
                        },
                    ],
                    "total_racks": 1120,
                    "all_compatible": True,
                    "blended_pue": 1.247,
                    "score": 91.4,
                    "trade_off_notes": [
                        "Keeps the majority share in lower-risk hyperscale demand.",
                        "Adds a limited AI/GPU tranche without pushing rack counts below threshold.",
                    ],
                },
                {
                    "rank": 2,
                    "allocations": [
                        {
                            "load_type": LoadType.HYPERSCALE.value,
                            "share_pct": 40.0,
                            "it_load_mw": 4.8,
                            "rack_count": 480,
                            "rack_density_kw": 10.0,
                        },
                        {
                            "load_type": LoadType.HPC.value,
                            "share_pct": 40.0,
                            "it_load_mw": 4.8,
                            "rack_count": 480,
                            "rack_density_kw": 10.0,
                        },
                        {
                            "load_type": LoadType.AI_GPU.value,
                            "share_pct": 20.0,
                            "it_load_mw": 2.4,
                            "rack_count": 160,
                            "rack_density_kw": 15.0,
                        },
                    ],
                    "total_racks": 1120,
                    "all_compatible": True,
                    "blended_pue": 1.254,
                    "score": 88.7,
                    "trade_off_notes": [
                        "Improves workload balance at a small efficiency penalty."
                    ],
                },
            ],
        },
    }


def make_green_energy_export_payload(
    *,
    result_key: str | None = None,
    pv_profile_source: str = "pvgis",
    site_id: str = "site-1",
    site_name: str = "Alpha Campus",
) -> dict:
    payload = {
        "result_key": result_key,
        "result": {
            "total_overhead_kwh": 12600000.0,
            "total_pv_generation_kwh": 5100000.0,
            "total_pv_to_overhead_kwh": 3700000.0,
            "total_pv_to_bess_kwh": 900000.0,
            "total_pv_curtailed_kwh": 500000.0,
            "total_bess_discharge_kwh": 760000.0,
            "total_fuel_cell_kwh": 420000.0,
            "total_grid_import_kwh": 11120000.0,
            "overhead_coverage_fraction": 0.3873,
            "renewable_fraction": 0.2191,
            "pv_self_consumption_fraction": 0.902,
            "bess_cycles_equivalent": 124.5,
            "co2_avoided_tonnes": 1234.6,
            "pv_capacity_kwp": 4800.0,
            "bess_capacity_kwh": 3000.0,
            "bess_roundtrip_efficiency": 0.875,
            "fuel_cell_capacity_kw": 700.0,
            "total_facility_kwh": 22380000.0,
            "total_it_kwh": 18000000.0,
            "site_name": site_name,
            "hours": 8760,
            "annual_pue": 1.243,
            "pue_source": "hourly",
            "nominal_it_mw": 12.0,
            "committed_it_mw": 10.8,
            "pv_profile_source": pv_profile_source,
            "pvgis_profile_key": "fixture-pvgis-profile" if pv_profile_source == "pvgis" else None,
        },
        "pv_profile_name": "fixture-pv.csv" if pv_profile_source == "manual" else None,
        "pvgis_profile": (
            {
                "site_id": site_id,
                "site_name": site_name,
                "profile_key": "fixture-pvgis-profile",
                "from_cache": True,
                "latitude": 45.0,
                "longitude": 9.0,
                "start_year": 2019,
                "end_year": 2023,
                "years_averaged": [2019, 2020, 2021, 2022, 2023],
                "pv_technology": "crystSi",
                "mounting_place": "free",
                "system_loss_pct": 14.0,
                "use_horizon": True,
                "optimal_angles": True,
                "surface_tilt_deg": None,
                "surface_azimuth_deg": None,
                "source": "PVGIS 5.3 representative-year profile",
                "radiation_database": "PVGIS-SARAH3",
                "elevation_m": 130.0,
                "pv_module_info": "Fixture module",
                "hours": 8760,
            }
            if pv_profile_source == "pvgis"
            else None
        ),
        "bess_initial_soc_kwh": 500.0,
        "grid_co2_kg_per_kwh": 0.256,
    }
    return payload


def workbook_sheet_rows(workbook_bytes: bytes, sheet_name: str) -> list[dict[str, object]]:
    sheet = load_workbook(BytesIO(workbook_bytes))[sheet_name]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, object]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        rows.append(
            {
                str(header[index]): row[index]
                for index in range(len(header))
                if header[index] is not None
            }
        )
    return rows


def workbook_summary_rows(workbook_bytes: bytes) -> dict[object, object]:
    return {
        row[0].value: row[1].value
        for row in load_workbook(BytesIO(workbook_bytes))["Summary"].iter_rows(
            min_row=2,
            max_col=2,
        )
        if row[0].value is not None
    }


def workbook_sheet_names(workbook_bytes: bytes) -> set[str]:
    return set(load_workbook(BytesIO(workbook_bytes)).sheetnames)


@pytest.fixture
def site_only_export_case() -> ExportTestCase:
    site_id = "site-only"
    site = make_site("Alpha Campus", latitude=45.0, longitude=9.0)
    primary = make_result(
        site_id,
        site.name,
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )
    return ExportTestCase(
        report_type="detailed",
        site_entries=[(site_id, site)],
        scenario_results=[primary],
        studied_site_ids=[site_id],
        primary_result_keys={site_id: get_result_selection_key(primary)},
    )


@pytest.fixture
def site_plus_grid_export_case() -> ExportTestCase:
    site_id = "site-grid"
    site = make_site(
        "Grid Campus",
        latitude=45.0,
        longitude=9.0,
        imported_geometry=ImportedGeometry(
            geometry_type="polygon",
            coordinates=[
                [45.0000, 9.0000],
                [45.0000, 9.0040],
                [45.0030, 9.0040],
                [45.0030, 9.0000],
                [45.0000, 9.0000],
            ],
        ),
    )
    primary = make_result(
        site_id,
        site.name,
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=91.0,
        annual_pue=1.24,
    )
    return ExportTestCase(
        report_type="detailed",
        site_entries=[(site_id, site)],
        scenario_results=[primary],
        studied_site_ids=[site_id],
        primary_result_keys={site_id: get_result_selection_key(primary)},
        grid_context_site_ids={site_id},
    )


@pytest.fixture
def site_plus_climate_export_case() -> ExportTestCase:
    site_id = "site-climate"
    site = make_site("Climate Campus", latitude=45.0, longitude=9.0)
    primary = make_result(
        site_id,
        site.name,
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.AIR_CHILLER_ECON,
        score=92.0,
        annual_pue=1.18,
    )
    temperatures, humidities = make_full_year_temperatures()
    return ExportTestCase(
        report_type="detailed",
        site_entries=[(site_id, site)],
        scenario_results=[primary],
        studied_site_ids=[site_id],
        primary_result_keys={site_id: get_result_selection_key(primary)},
        weather_by_site={
            site_id: make_weather_export_payload(
                temperatures=temperatures,
                humidities=humidities,
                years_averaged=[2024],
                include_coordinates=True,
            )
        },
    )


@pytest.fixture
def full_analysis_export_case() -> ExportTestCase:
    site_id = "site-full"
    site = make_site(
        "Full Analysis Campus",
        latitude=45.0,
        longitude=9.0,
        power_confirmed=True,
        imported_geometry=ImportedGeometry(
            geometry_type="polygon",
            coordinates=[
                [45.0000, 9.0000],
                [45.0000, 9.0040],
                [45.0030, 9.0040],
                [45.0030, 9.0000],
                [45.0000, 9.0000],
            ],
        ),
    )
    primary = make_result(
        site_id,
        site.name,
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
        overtemperature_hours=48,
        it_capacity_worst_mw=9.4,
        it_capacity_p99_mw=10.8,
        it_capacity_p90_mw=11.1,
        it_capacity_mean_mw=11.4,
        it_capacity_best_mw=12.2,
    )
    alternative = make_result(
        site_id,
        site.name,
        load_type=LoadType.EDGE_TELCO,
        cooling_type=CoolingType.CRAC_DX,
        score=71.0,
        annual_pue=1.44,
    )
    primary_key = get_result_selection_key(primary)
    temperatures, humidities = make_full_year_temperatures()
    return ExportTestCase(
        report_type="detailed",
        site_entries=[(site_id, site)],
        scenario_results=[primary, alternative],
        studied_site_ids=[site_id],
        primary_result_keys={site_id: primary_key},
        grid_context_site_ids={site_id},
        weather_by_site={
            site_id: make_weather_export_payload(
                temperatures=temperatures,
                humidities=humidities,
                years_averaged=[2024],
                include_coordinates=True,
            )
        },
        load_mix_results={site_id: make_load_mix_export_payload(result_key=primary_key)},
        green_energy_results={
            site_id: make_green_energy_export_payload(
                result_key=primary_key,
                site_id=site_id,
                site_name=site.name,
            )
        },
    )


@pytest.mark.parametrize(
    ("case_fixture", "expected"),
    [
        (
            "site_only_export_case",
            {
                "grid": False,
                "climate": False,
                "load_mix": False,
                "green_energy": False,
                "full_matrix_rows": 1,
            },
        ),
        (
            "site_plus_grid_export_case",
            {
                "grid": True,
                "climate": False,
                "load_mix": False,
                "green_energy": False,
                "full_matrix_rows": 1,
            },
        ),
        (
            "site_plus_climate_export_case",
            {
                "grid": False,
                "climate": True,
                "load_mix": False,
                "green_energy": False,
                "full_matrix_rows": 1,
            },
        ),
        (
            "full_analysis_export_case",
            {
                "grid": True,
                "climate": True,
                "load_mix": True,
                "green_energy": True,
                "full_matrix_rows": 2,
            },
        ),
    ],
)
def test_build_report_bundle_matches_export_shape_fixture(
    case_fixture,
    expected,
    request,
    monkeypatch,
):
    case = request.getfixturevalue(case_fixture)
    case.install_optional_inputs(monkeypatch, make_local_temp_dir())

    bundle = build_report_bundle(**case.export_kwargs())

    site_bundle = bundle["report_bundle"]["studied_sites"][0]
    chapters = site_bundle["chapters"]
    availability = bundle["report_bundle"]["analysis_availability"]

    assert site_bundle["site_id"] == case.studied_site_ids[0]
    assert chapters["selected_scenario"]["available"] is True
    assert chapters["deep_dive"]["available"] is True
    assert chapters["grid_context"]["included"] is expected["grid"]
    assert chapters["climate"]["included"] is expected["climate"]
    assert chapters["load_mix"]["included"] is expected["load_mix"]
    assert chapters["green_energy"]["included"] is expected["green_energy"]
    assert site_bundle["grid_context"]["status"] == (
        "available" if expected["grid"] else "missing"
    )
    assert site_bundle["climate"]["status"] == (
        "available" if expected["climate"] else "missing"
    )
    assert availability["grid_context_available_site_count"] == int(expected["grid"])
    assert availability["climate_available_site_count"] == int(expected["climate"])
    assert availability["load_mix_available_site_count"] == int(expected["load_mix"])
    assert availability["green_energy_available_site_count"] == int(
        expected["green_energy"]
    )
    assert bundle["summary"]["primary_result_count"] == 1
    assert bundle["summary"]["scenario_count"] == expected["full_matrix_rows"]
    assert bundle["summary"]["available_scenario_count"] == expected["full_matrix_rows"]


@pytest.mark.parametrize(
    ("case_fixture", "present_sections", "absent_sections"),
    [
        (
            "site_only_export_case",
            set(),
            {
                "Grid Infrastructure",
                "Weather &amp; Climate",
                "Load Mix Analysis",
                "Green Energy Dispatch",
            },
        ),
        (
            "site_plus_grid_export_case",
            {"Grid Infrastructure", "Grid Context Map"},
            {
                "Weather &amp; Climate",
                "Load Mix Analysis",
                "Green Energy Dispatch",
            },
        ),
        (
            "site_plus_climate_export_case",
            {"Weather &amp; Climate", "Monthly Temperature Profile", "Free Cooling Hours by Topology"},
            {
                "Grid Infrastructure",
                "Load Mix Analysis",
                "Green Energy Dispatch",
            },
        ),
        (
            "full_analysis_export_case",
            {
                "Grid Infrastructure",
                "Weather &amp; Climate",
                "Load Mix Analysis",
                "Green Energy Dispatch",
            },
            set(),
        ),
    ],
)
def test_render_report_html_matches_export_shape_fixture_sections(
    case_fixture,
    present_sections,
    absent_sections,
    request,
    monkeypatch,
):
    case = request.getfixturevalue(case_fixture)
    case.install_optional_inputs(monkeypatch, make_local_temp_dir())

    html = render_report_html(**case.export_kwargs())

    assert case.site_entries[0][1].name in html
    assert "Site Specifications" in html
    assert "Scenario Results" in html
    for section_title in present_sections:
        assert section_title in html
    for section_title in absent_sections:
        assert section_title not in html


@pytest.mark.parametrize(
    ("case_fixture", "present_sheets", "absent_sheets", "expected_counts"),
    [
        (
            "site_only_export_case",
            set(),
            {
                "Grid Summary",
                "Grid Assets",
                "Climate Summary",
                "Climate Free Cooling",
                "Load Mix",
                "Load Mix Candidates",
                "Green Energy",
            },
            {"grid": 0, "climate": 0, "load_mix": 0, "green_energy": 0, "full_matrix": 1, "displayed": 1},
        ),
        (
            "site_plus_grid_export_case",
            {"Grid Summary", "Grid Assets"},
            {
                "Climate Summary",
                "Climate Free Cooling",
                "Load Mix",
                "Load Mix Candidates",
                "Green Energy",
            },
            {"grid": 1, "climate": 0, "load_mix": 0, "green_energy": 0, "full_matrix": 1, "displayed": 1},
        ),
        (
            "site_plus_climate_export_case",
            {"Climate Summary", "Climate Free Cooling"},
            {
                "Grid Summary",
                "Grid Assets",
                "Load Mix",
                "Load Mix Candidates",
                "Green Energy",
            },
            {"grid": 0, "climate": 1, "load_mix": 0, "green_energy": 0, "full_matrix": 1, "displayed": 1},
        ),
        (
            "full_analysis_export_case",
            {
                "Grid Summary",
                "Grid Assets",
                "Climate Summary",
                "Climate Free Cooling",
                "Load Mix",
                "Load Mix Candidates",
                "Green Energy",
                "Appx Summary",
                "Appx Tables",
            },
            set(),
            {"grid": 1, "climate": 1, "load_mix": 1, "green_energy": 1, "full_matrix": 2, "displayed": 2},
        ),
    ],
)
def test_build_excel_bytes_matches_export_shape_fixture_sheets(
    case_fixture,
    present_sheets,
    absent_sheets,
    expected_counts,
    request,
    monkeypatch,
):
    case = request.getfixturevalue(case_fixture)
    case.install_optional_inputs(monkeypatch, make_local_temp_dir())

    workbook = build_excel_bytes(**case.export_kwargs())

    sheet_names = workbook_sheet_names(workbook)
    summary_rows = workbook_summary_rows(workbook)

    assert {"Summary", "Metadata", "Sites", "Scenarios", "Primary Results"}.issubset(
        sheet_names
    )
    assert present_sheets.issubset(sheet_names)
    assert absent_sheets.isdisjoint(sheet_names)
    assert summary_rows["Scenario Results Included"] == expected_counts["displayed"]
    assert summary_rows["Full Scenario Matrix Rows"] == expected_counts["full_matrix"]
    assert summary_rows["Grid Context Sites Available"] == expected_counts["grid"]
    assert summary_rows["Climate Sites Available"] == expected_counts["climate"]
    assert summary_rows["Load Mix Sites Available"] == expected_counts["load_mix"]
    assert summary_rows["Green Energy Sites Available"] == expected_counts["green_energy"]


def test_full_analysis_export_keeps_primary_story_and_full_matrix_depth(
    full_analysis_export_case,
    monkeypatch,
):
    full_analysis_export_case.install_optional_inputs(monkeypatch, make_local_temp_dir())

    html = render_report_html(**full_analysis_export_case.export_kwargs())
    workbook = build_excel_bytes(**full_analysis_export_case.export_kwargs())

    assert CoolingType.WATER_CHILLER_ECON.value in html
    assert LoadType.HYPERSCALE.value in html
    # With include_all_scenarios=True (default), all scenarios appear in the HTML comparison table
    assert LoadType.EDGE_TELCO.value in html
    assert {
        row["Cooling Type"] for row in workbook_sheet_rows(workbook, "Scenarios")
    } == {
        CoolingType.WATER_CHILLER_ECON.value,
        CoolingType.CRAC_DX.value,
    }
    assert workbook_summary_rows(workbook)["Scenario Results Included"] == 2
    assert workbook_summary_rows(workbook)["Full Scenario Matrix Rows"] == 2


def test_render_report_html_with_site_only():
    site = make_site("Test Site")

    html = render_report_html(
        report_type="executive",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[],
    )

    assert "Executive Feasibility Summary" in html
    assert "Test Site" in html


def test_build_excel_bytes_with_site_only():
    site = make_site("Test Site")

    workbook = build_excel_bytes(
        report_type="executive",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[],
    )

    assert workbook[:2] == b"PK"


def test_build_excel_bytes_uses_full_studied_site_scenario_matrix():
    site_one = make_site("Alpha Campus")
    site_two = make_site("Beta Campus")
    site_three = make_site("Gamma Campus")

    site_one_primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )
    site_one_alternative = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.EDGE_TELCO,
        cooling_type=CoolingType.CRAC_DX,
        score=71.0,
        annual_pue=1.44,
    )
    site_two_primary = make_result(
        "site-2",
        "Beta Campus",
        load_type=LoadType.HPC,
        cooling_type=CoolingType.DLC,
        score=89.0,
        annual_pue=1.18,
    )
    site_three_result = make_result(
        "site-3",
        "Gamma Campus",
        load_type=LoadType.COLOCATION_STANDARD,
        cooling_type=CoolingType.AIR_CHILLER_ECON,
        score=82.0,
        annual_pue=1.27,
    )

    workbook = build_excel_bytes(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[
            ("site-1", site_one),
            ("site-2", site_two),
            ("site-3", site_three),
        ],
        scenario_results=[
            site_one_primary,
            site_one_alternative,
            site_two_primary,
            site_three_result,
        ],
        studied_site_ids=["site-1", "site-2"],
        primary_result_keys={
            "site-1": get_result_selection_key(site_one_primary),
            "site-2": get_result_selection_key(site_two_primary),
        },
    )

    summary_rows = workbook_summary_rows(workbook)
    scenario_rows = workbook_sheet_rows(workbook, "Scenarios")
    primary_rows = workbook_sheet_rows(workbook, "Primary Results")
    site_rows = workbook_sheet_rows(workbook, "Sites")
    metadata_rows = workbook_sheet_rows(workbook, "Metadata")

    # With include_all_scenarios=True (default), all studied-site results are displayed
    assert summary_rows["Scenario Results Included"] == 3
    assert summary_rows["Full Scenario Matrix Rows"] == 3
    assert len(primary_rows) == 2
    assert len(scenario_rows) == 3
    assert {row["Site ID"] for row in scenario_rows} == {"site-1", "site-2"}
    assert any(row["Cooling Type"] == CoolingType.CRAC_DX.value for row in scenario_rows)
    assert {row["Site ID"] for row in site_rows} == {"site-1", "site-2"}
    assert any(
        row["Key"] == "Studied site IDs" and row["Value"] == "site-1; site-2"
        for row in metadata_rows
    )


def test_build_excel_bytes_adds_optional_depth_sheets(monkeypatch):
    site = make_site(
        "Alpha Campus",
        latitude=45.0,
        longitude=9.0,
        available_power_mw=20.0,
        power_confirmed=True,
    )
    primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
        it_capacity_worst_mw=9.4,
        it_capacity_p99_mw=10.8,
        it_capacity_p90_mw=11.1,
        it_capacity_mean_mw=11.4,
        it_capacity_best_mw=12.2,
    )
    primary_key = get_result_selection_key(primary)

    temperatures, humidities = make_full_year_temperatures()
    temp_dir = make_local_temp_dir()
    write_grid_context_cache(temp_dir, "site-1", "Alpha Campus")
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(
        report_data,
        "get_weather",
        lambda site_id: {
            "temperatures": temperatures,
            "humidities": humidities,
            "source": "Fixture Weather",
            "source_type": "manual_upload",
            "hours": len(temperatures),
            "years_averaged": [2024],
            "original_filename": "fixture-weather.csv",
            "latitude": 45.0,
            "longitude": 9.0,
        },
    )

    workbook = build_excel_bytes(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": primary_key},
        load_mix_results={"site-1": make_load_mix_export_payload(result_key=primary_key)},
        green_energy_results={
            "site-1": make_green_energy_export_payload(result_key=primary_key)
        },
    )

    loaded = load_workbook(BytesIO(workbook))

    assert {
        "Grid Summary",
        "Grid Assets",
        "Climate Summary",
        "Climate Free Cooling",
        "Load Mix",
        "Load Mix Candidates",
        "Green Energy",
        "Appx Summary",
        "Appx Tables",
    }.issubset(set(loaded.sheetnames))
    assert workbook_sheet_rows(workbook, "Grid Summary")[0]["Status"] == "available"
    assert any(
        row["Block Key"] == "backup_comparison"
        for row in workbook_sheet_rows(workbook, "Appx Summary")
    )
    assert any(
        row["Column Label"] == "Technology"
        for row in workbook_sheet_rows(workbook, "Appx Tables")
    )


def test_build_excel_bytes_omits_optional_analysis_sheets_when_missing(monkeypatch):
    site = make_site("Alpha Campus")
    primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )

    temp_dir = make_local_temp_dir()
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(report_data, "get_weather", lambda site_id: None)

    workbook = build_excel_bytes(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(primary)},
    )

    loaded = load_workbook(BytesIO(workbook))

    assert "Metadata" in loaded.sheetnames
    assert "Scenarios" in loaded.sheetnames
    assert "Grid Summary" not in loaded.sheetnames
    assert "Climate Summary" not in loaded.sheetnames
    assert "Load Mix" not in loaded.sheetnames
    assert "Green Energy" not in loaded.sheetnames


def test_build_report_context_filters_to_selected_studied_sites():
    site_one = make_site("Alpha Campus")
    site_two = make_site("Beta Campus")
    site_three = make_site("Gamma Campus")

    results = [
        make_result("site-1", "Alpha Campus", score=91.0),
        make_result("site-2", "Beta Campus", score=88.0),
        make_result("site-3", "Gamma Campus", score=85.0),
    ]

    context = build_report_context(
        report_type="executive",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[
            ("site-1", site_one),
            ("site-2", site_two),
            ("site-3", site_three),
        ],
        scenario_results=results,
        studied_site_ids=["site-1", "site-2"],
    )

    assert [section["site_id"] for section in context["site_sections"]] == ["site-1", "site-2"]
    assert {result.site_id for result in context["ranked_results"]} == {"site-1", "site-2"}
    assert context["summary"]["site_count"] == 2


def test_build_report_context_filters_to_selected_primary_results():
    site_one = make_site("Alpha Campus")
    site_two = make_site("Beta Campus")

    site_one_primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )
    site_one_alternative = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.EDGE_TELCO,
        cooling_type=CoolingType.CRAC_DX,
        score=71.0,
        annual_pue=1.44,
    )
    site_two_primary = make_result(
        "site-2",
        "Beta Campus",
        load_type=LoadType.HPC,
        cooling_type=CoolingType.DLC,
        score=89.0,
        annual_pue=1.18,
    )

    context = build_report_context(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site_one), ("site-2", site_two)],
        scenario_results=[
            site_one_primary,
            site_one_alternative,
            site_two_primary,
        ],
        studied_site_ids=["site-1", "site-2"],
        primary_result_keys={
            "site-1": get_result_selection_key(site_one_primary),
            "site-2": get_result_selection_key(site_two_primary),
        },
    )

    # With include_all_scenarios=True (default), all results for studied sites are included
    assert len(context["ranked_results"]) == 3
    assert {result.site_id for result in context["ranked_results"]} == {"site-1", "site-2"}
    assert context["site_sections"][0]["result_count"] == 2
    assert (
        context["site_sections"][0]["selected_result"].scenario.cooling_type
        == CoolingType.WATER_CHILLER_ECON
    )


def test_build_report_bundle_normalizes_primary_and_alternatives(monkeypatch):
    site = make_site("Alpha Campus")
    primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )
    alternative = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.EDGE_TELCO,
        cooling_type=CoolingType.CRAC_DX,
        score=71.0,
        annual_pue=1.44,
    )

    temp_dir = make_local_temp_dir()
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(report_data, "get_weather", lambda site_id: None)

    bundle = build_report_bundle(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary, alternative],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(primary)},
    )

    site_bundle = bundle["report_bundle"]["studied_sites"][0]

    assert site_bundle["results"]["primary"]["result_key"] == get_result_selection_key(primary)
    assert site_bundle["results"]["primary"]["scenario"]["cooling_type"] == (
        CoolingType.WATER_CHILLER_ECON.value
    )
    assert len(site_bundle["results"]["alternatives"]) == 1
    assert site_bundle["results"]["alternatives"][0]["result_key"] == get_result_selection_key(
        alternative
    )
    assert bundle["report_bundle"]["resolved_primary_result_keys"] == {
        "site-1": get_result_selection_key(primary)
    }
    assert bundle["study_scope"]["selected_primary_result_count"] == 1


def test_build_report_bundle_includes_cached_grid_context(monkeypatch):
    site = make_site("Alpha Campus")
    primary = make_result("site-1", "Alpha Campus", score=92.0)

    temp_dir = make_local_temp_dir()
    write_grid_context_cache(temp_dir, "site-1", "Alpha Campus")
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(report_data, "get_weather", lambda site_id: None)

    bundle = build_report_bundle(
        report_type="executive",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(primary)},
    )

    grid_context = bundle["report_bundle"]["studied_sites"][0]["grid_context"]

    assert grid_context["status"] == "available"
    assert grid_context["selected"]["summary"]["radius_km"] == pytest.approx(5.0)
    assert grid_context["selected"]["asset_count"] == 1
    assert (
        bundle["report_bundle"]["analysis_availability"]["grid_context_available_site_count"]
        == 1
    )


def test_build_report_bundle_includes_climate_analysis_from_cached_weather(
    monkeypatch,
):
    site = make_site("Alpha Campus")
    primary = make_result(
        "site-1",
        "Alpha Campus",
        cooling_type=CoolingType.AIR_CHILLER_ECON,
        score=92.0,
    )

    temp_dir = make_local_temp_dir()
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(
        report_data,
        "get_weather",
        lambda site_id: {
            "temperatures": [10.0, 12.0, 14.0, 16.0],
            "humidities": [50.0, 52.0, 48.0, 47.0],
            "source": "Fixture Weather",
            "source_type": "manual_upload",
            "hours": 4,
            "years_averaged": [],
            "original_filename": "fixture-weather.csv",
        },
    )

    bundle = build_report_bundle(
        report_type="executive",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(primary)},
    )

    climate = bundle["report_bundle"]["studied_sites"][0]["climate"]

    assert climate["status"] == "available"
    assert climate["weather_status"]["source_type"] == "manual_upload"
    assert climate["analysis"]["temperature_stats"]["count"] == 4
    assert climate["analysis"]["cooling_types_analyzed"] == [
        CoolingType.AIR_CHILLER_ECON.value
    ]
    assert (
        bundle["report_bundle"]["analysis_availability"]["climate_available_site_count"]
        == 1
    )


def test_build_report_bundle_shapes_core_chapter_context(monkeypatch):
    site = make_site("Alpha Campus")
    primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )

    temp_dir = make_local_temp_dir()
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(report_data, "get_weather", lambda site_id: None)

    bundle = build_report_bundle(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(primary)},
    )

    chapters = bundle["report_bundle"]["studied_sites"][0]["chapters"]

    assert chapters["site_specifics"]["title"] == "Site Specifics and Properties"
    assert chapters["grid_context"]["included"] is False
    assert chapters["climate"]["included"] is False
    assert chapters["selected_scenario"]["available"] is True
    assert chapters["deep_dive"]["available"] is True
    assert chapters["load_mix"]["included"] is False
    assert chapters["green_energy"]["included"] is False
    assert bundle["report"]["narrative_policy"]["mode"] == "structured_guardrail_v1"
    assert bundle["report"]["narrative"]["available"] is True
    assert chapters["site_specifics"]["narrative"]["available"] is True
    assert chapters["selected_scenario"]["narrative"]["available"] is True
    assert chapters["deep_dive"]["narrative"]["available"] is True
    assert any(
        item["value"] == CoolingType.WATER_CHILLER_ECON.value
        for item in chapters["selected_scenario"]["summary_items"]
    )
    assert any(
        item["label"] == "Committed IT capacity"
        for item in chapters["deep_dive"]["headline_metrics"]
    )
    assert chapters["deep_dive"]["advanced_block_count"] == 4
    assert {
        block["key"] for block in chapters["deep_dive"]["advanced_blocks"]
    } == {
        "expansion_advisory",
        "footprint",
        "backup_comparison",
        "sensitivity",
    }


def test_build_report_bundle_shapes_milestone_five_advanced_blocks(monkeypatch):
    site = make_site("Alpha Campus", power_confirmed=True)
    primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
        overtemperature_hours=48,
        it_capacity_worst_mw=9.4,
        it_capacity_p99_mw=10.8,
        it_capacity_p90_mw=11.1,
        it_capacity_mean_mw=11.4,
        it_capacity_best_mw=12.2,
    )

    temperatures, humidities = make_full_year_temperatures()
    temp_dir = make_local_temp_dir()
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(
        report_data,
        "get_weather",
        lambda site_id: {
            "temperatures": temperatures,
            "humidities": humidities,
            "source": "Fixture Weather",
            "source_type": "manual_upload",
            "hours": len(temperatures),
            "years_averaged": [2024],
            "original_filename": "fixture-weather.csv",
        },
    )

    bundle = build_report_bundle(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(primary)},
    )

    blocks = bundle["report_bundle"]["studied_sites"][0]["chapters"]["deep_dive"]["advanced_blocks"]
    assert [block["key"] for block in blocks] == [
        "pue_decomposition",
        "hourly_profiles",
        "it_capacity_spectrum",
        "expansion_advisory",
        "firm_capacity",
        "footprint",
        "backup_comparison",
        "sensitivity",
        "break_even",
    ]
    assert any(block["title"] == "PUE Decomposition" for block in blocks)
    assert any(block["title"] == "Firm Capacity" for block in blocks)
    assert all(block["tables"] for block in blocks)


def test_build_report_bundle_shapes_milestone_six_optional_chapters(monkeypatch):
    site = make_site("Alpha Campus")
    primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )
    primary_key = get_result_selection_key(primary)

    temp_dir = make_local_temp_dir()
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(report_data, "get_weather", lambda site_id: None)

    bundle = build_report_bundle(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": primary_key},
        load_mix_results={"site-1": make_load_mix_export_payload(result_key=primary_key)},
        green_energy_results={
            "site-1": make_green_energy_export_payload(result_key=primary_key)
        },
    )

    site_bundle = bundle["report_bundle"]["studied_sites"][0]
    chapters = site_bundle["chapters"]

    assert site_bundle["load_mix"]["status"] == "available"
    assert site_bundle["green_energy"]["status"] == "available"
    assert chapters["load_mix"]["included"] is True
    assert chapters["load_mix"]["has_candidates"] is True
    assert chapters["load_mix"]["narrative"]["available"] is True
    assert chapters["load_mix"]["top_candidate_table"]["rows"][0]["load_type"] == (
        LoadType.HYPERSCALE.value
    )
    assert chapters["green_energy"]["included"] is True
    assert chapters["green_energy"]["narrative"]["available"] is True
    assert any(
        item["label"] == "Renewable fraction"
        for item in chapters["green_energy"]["headline_items"]
    )
    assert (
        bundle["report_bundle"]["analysis_availability"]["load_mix_available_site_count"]
        == 1
    )
    assert (
        bundle["report_bundle"]["analysis_availability"]["green_energy_available_site_count"]
        == 1
    )


def test_build_report_bundle_omits_stale_milestone_six_payloads(monkeypatch):
    site = make_site("Alpha Campus")
    primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )

    temp_dir = make_local_temp_dir()
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(report_data, "get_weather", lambda site_id: None)

    bundle = build_report_bundle(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(primary)},
        load_mix_results={"site-1": make_load_mix_export_payload(result_key="stale-key")},
        green_energy_results={
            "site-1": make_green_energy_export_payload(result_key="stale-key")
        },
    )

    site_bundle = bundle["report_bundle"]["studied_sites"][0]

    assert site_bundle["load_mix"]["status"] == "missing"
    assert site_bundle["green_energy"]["status"] == "missing"
    assert site_bundle["chapters"]["load_mix"]["included"] is False
    assert site_bundle["chapters"]["green_energy"]["included"] is False


def test_build_report_bundle_shapes_export_safe_visuals(monkeypatch):
    site = make_site(
        "Alpha Campus",
        latitude=45.0,
        longitude=9.0,
        imported_geometry=ImportedGeometry(
            geometry_type="polygon",
            coordinates=[
                [45.0000, 9.0000],
                [45.0000, 9.0040],
                [45.0030, 9.0040],
                [45.0030, 9.0000],
                [45.0000, 9.0000],
            ],
        ),
    )
    primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )

    temperatures, humidities = make_full_year_temperatures()
    temp_dir = make_local_temp_dir()
    write_grid_context_cache(temp_dir, "site-1", "Alpha Campus")
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(
        report_data,
        "get_weather",
        lambda site_id: {
            "temperatures": temperatures,
            "humidities": humidities,
            "source": "Fixture Weather",
            "source_type": "manual_upload",
            "hours": len(temperatures),
            "years_averaged": [2024],
            "original_filename": "fixture-weather.csv",
            "latitude": 45.0,
            "longitude": 9.0,
        },
    )

    bundle = build_report_bundle(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(primary)},
    )

    chapters = bundle["report_bundle"]["studied_sites"][0]["chapters"]

    assert chapters["site_specifics"]["map_visual"]["available"] is True
    assert "<svg" in chapters["site_specifics"]["map_visual"]["svg_markup"]
    assert chapters["grid_context"]["map_visual"]["available"] is True
    assert "<svg" in chapters["grid_context"]["map_visual"]["svg_markup"]
    assert chapters["climate"]["monthly_chart_visual"]["available"] is True
    assert chapters["climate"]["free_cooling_chart_visual"]["available"] is True
    assert "<svg" in chapters["climate"]["monthly_chart_visual"]["svg_markup"]
    assert "<svg" in chapters["climate"]["free_cooling_chart_visual"]["svg_markup"]


def test_export_still_works_when_optional_analyses_are_missing(monkeypatch):
    site = make_site("Alpha Campus")
    primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )

    temp_dir = make_local_temp_dir()
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(report_data, "get_weather", lambda site_id: None)

    context = build_report_bundle(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(primary)},
    )
    html = render_report_html(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(primary)},
    )
    workbook = build_excel_bytes(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(primary)},
    )

    site_bundle = context["report_bundle"]["studied_sites"][0]

    assert site_bundle["grid_context"]["status"] == "missing"
    assert site_bundle["climate"]["status"] == "missing"
    assert "Alpha Campus" in html
    assert "Grid Infrastructure" not in html
    assert "Weather &amp; Climate" not in html
    assert "Site Specifications" in html
    assert "Scenario Results" in html
    assert "Load Mix Analysis" not in html
    assert "Green Energy Dispatch" not in html
    assert workbook[:2] == b"PK"


def test_render_report_html_omits_missing_hourly_advanced_blocks(monkeypatch):
    site = make_site("Alpha Campus")
    primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
        it_capacity_worst_mw=9.4,
        it_capacity_p99_mw=10.8,
        it_capacity_p90_mw=11.1,
        it_capacity_mean_mw=11.4,
        it_capacity_best_mw=12.2,
    )

    temp_dir = make_local_temp_dir()
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(report_data, "get_weather", lambda site_id: None)

    html = render_report_html(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(primary)},
    )

    assert "PUE Decomposition" not in html
    assert "Hourly Profiles" not in html
    assert "Firm Capacity" not in html
    assert "IT Capacity Spectrum" in html
    assert "Expansion Advisory" in html
    assert "Footprint" in html
    assert "Backup Comparison" in html
    assert "Sensitivity" in html
    assert "Break-Even" in html


def test_render_report_html_includes_core_chapters_when_optional_analyses_exist(
    monkeypatch,
):
    site = make_site(
        "Alpha Campus",
        latitude=45.0,
        longitude=9.0,
        imported_geometry=ImportedGeometry(
            geometry_type="polygon",
            coordinates=[
                [45.0000, 9.0000],
                [45.0000, 9.0040],
                [45.0030, 9.0040],
                [45.0030, 9.0000],
                [45.0000, 9.0000],
            ],
        ),
    )
    primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )

    temperatures, humidities = make_full_year_temperatures()
    temp_dir = make_local_temp_dir()
    write_grid_context_cache(temp_dir, "site-1", "Alpha Campus")
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(
        report_data,
        "get_weather",
        lambda site_id: {
            "temperatures": temperatures,
            "humidities": humidities,
            "source": "Fixture Weather",
            "source_type": "manual_upload",
            "hours": len(temperatures),
            "years_averaged": [2024],
            "original_filename": "fixture-weather.csv",
            "latitude": 45.0,
            "longitude": 9.0,
        },
    )

    html = render_report_html(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(primary)},
    )

    assert "Site Specifications" in html
    assert "Site Map" in html
    assert "Grid Infrastructure" in html
    assert "Grid Context Map" in html
    assert "Weather &amp; Climate" in html
    assert "Monthly Temperature Profile" in html
    assert "Cooling Topology Suitability" in html
    assert "Scenario Results" in html
    assert "Mapped Infrastructure Assets" in html
    assert "Free Cooling Analysis" in html
    assert CoolingType.WATER_CHILLER_ECON.value in html


def test_render_report_html_includes_milestone_six_optional_chapters(monkeypatch):
    site = make_site("Alpha Campus")
    primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )
    primary_key = get_result_selection_key(primary)

    temp_dir = make_local_temp_dir()
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(report_data, "get_weather", lambda site_id: None)

    html = render_report_html(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": primary_key},
        load_mix_results={"site-1": make_load_mix_export_payload(result_key=primary_key)},
        green_energy_results={
            "site-1": make_green_energy_export_payload(result_key=primary_key)
        },
    )

    assert "Load Mix Analysis" in html
    assert "Top candidate mix" in html
    assert "Ranked candidate overview" in html
    assert "Green Energy Dispatch" in html
    assert "Annual energy breakdown" in html
    assert "Cached PVGIS normalized profile" in html
    assert "Renewable fraction" in html


def test_render_report_html_includes_guardrailed_narratives(monkeypatch):
    site = make_site("Alpha Campus")
    primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )
    primary_key = get_result_selection_key(primary)

    temp_dir = make_local_temp_dir()
    monkeypatch.setattr(report_data, "GRID_CONTEXT_DIR", temp_dir)
    monkeypatch.setattr(report_data, "get_weather", lambda site_id: None)

    html = render_report_html(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site)],
        scenario_results=[primary],
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": primary_key},
        load_mix_results={"site-1": make_load_mix_export_payload(result_key=primary_key)},
        green_energy_results={
            "site-1": make_green_energy_export_payload(result_key=primary_key)
        },
    )

    assert html.count("Structured basis:") >= 1
    assert "The optimizer evaluated 6 candidate mix(es)" in html
    assert "The dispatch run reaches 21.9% renewable fraction" in html


def test_render_report_html_and_excel_include_layout_mode_and_filtered_results():
    site_one = make_site("Alpha Campus")
    site_one_primary = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.HYPERSCALE,
        cooling_type=CoolingType.WATER_CHILLER_ECON,
        score=94.0,
        annual_pue=1.21,
    )
    site_one_alternative = make_result(
        "site-1",
        "Alpha Campus",
        load_type=LoadType.EDGE_TELCO,
        cooling_type=CoolingType.CRAC_DX,
        score=71.0,
        annual_pue=1.44,
    )

    html = render_report_html(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site_one)],
        scenario_results=[site_one_primary, site_one_alternative],
        layout_mode="report_a4_portrait",
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(site_one_primary)},
        include_all_scenarios=False,
    )

    html_presentation = render_report_html(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site_one)],
        scenario_results=[site_one_primary, site_one_alternative],
        layout_mode="presentation_16_9",
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(site_one_primary)},
        include_all_scenarios=False,
    )

    workbook = build_excel_bytes(
        report_type="detailed",
        primary_color="#1a365d",
        secondary_color="#2b6cb0",
        font_family="Inter, sans-serif",
        logo_url=None,
        site_entries=[("site-1", site_one)],
        scenario_results=[site_one_primary, site_one_alternative],
        layout_mode="report_a4_portrait",
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": get_result_selection_key(site_one_primary)},
        include_all_scenarios=False,
    )

    summary_sheet = load_workbook(BytesIO(workbook))["Summary"]
    summary_rows = {
        row[0].value: row[1].value
        for row in summary_sheet.iter_rows(min_row=2, max_col=2)
        if row[0].value is not None
    }

    assert "Report A4 Portrait" in html
    assert 'class="report-shell layout-report_a4_portrait report-type-detailed"' in html
    assert "size: A4 portrait;" in html
    assert 'class="report-shell layout-presentation_16_9 report-type-detailed"' in html_presentation
    assert "size: 297mm 167mm;" in html_presentation
    assert "body.layout-report_a4_portrait .chapter-head" in html
    assert CoolingType.WATER_CHILLER_ECON.value in html
    assert CoolingType.CRAC_DX.value not in html
    assert summary_rows["Layout Mode"] == "Report A4 Portrait"
    assert summary_rows["Scenario Results Included"] == 1


def test_export_html_endpoint_passes_selection_contract_to_renderer(monkeypatch):
    site = make_site("Alpha Campus")
    selected_result = make_result("site-1", "Alpha Campus", score=90.0)
    alternate_result = make_result(
        "site-1",
        "Alpha Campus",
        cooling_type=CoolingType.CRAC_DX,
        score=70.0,
    )
    captured: dict[str, object] = {}

    monkeypatch.setattr(routes_export, "get_site", lambda site_id: (site_id, site))

    def fake_render_report_html(**kwargs):
        captured.update(kwargs)
        return "<html>ok</html>"

    monkeypatch.setattr(routes_export, "render_report_html", fake_render_report_html)

    selected_key = get_result_selection_key(selected_result)
    config = routes_export.ReportConfig(
        report_type="executive",
        studied_site_ids=["site-1"],
        primary_result_keys={"site-1": selected_key},
        scenario_results=[
            selected_result.model_dump(mode="json"),
            alternate_result.model_dump(mode="json"),
        ],
        load_mix_results={"site-1": make_load_mix_export_payload(result_key=selected_key)},
        green_energy_results={
            "site-1": make_green_energy_export_payload(result_key=selected_key)
        },
        layout_mode="presentation_16_9",
    )

    response = asyncio.run(routes_export.export_html_endpoint(config))

    assert response.body == b"<html>ok</html>"
    assert captured["studied_site_ids"] == ["site-1"]
    assert captured["primary_result_keys"] == {"site-1": selected_key}
    assert captured["layout_mode"] == "presentation_16_9"
    assert captured["load_mix_results"] == {
        "site-1": make_load_mix_export_payload(result_key=selected_key)
    }
    assert captured["green_energy_results"] == {
        "site-1": make_green_energy_export_payload(result_key=selected_key)
    }
    assert [site_id for site_id, _ in captured["site_entries"]] == ["site-1"]


def test_export_html_endpoint_rejects_missing_primary_result_selection(monkeypatch):
    site = make_site("Alpha Campus")
    selected_result = make_result("site-1", "Alpha Campus", score=90.0)

    monkeypatch.setattr(routes_export, "get_site", lambda site_id: (site_id, site))

    config = routes_export.ReportConfig(
        report_type="executive",
        studied_site_ids=["site-1"],
        primary_result_keys={},
        scenario_results=[selected_result.model_dump(mode="json")],
        layout_mode="presentation_16_9",
    )

    with pytest.raises(routes_export.HTTPException) as excinfo:
        asyncio.run(routes_export.export_html_endpoint(config))

    assert excinfo.value.status_code == 400
    assert "Missing primary result selection" in excinfo.value.detail


@pytest.mark.parametrize(
    ("primary_result_keys", "expected_detail"),
    [
        (
            {"site-2": "unexpected-key"},
            "Primary result keys were provided for unselected studied sites",
        ),
        (
            {"site-1": "stale-key"},
            "Primary result selection did not match the current batch results",
        ),
    ],
)
def test_export_html_endpoint_rejects_invalid_primary_result_contract(
    monkeypatch,
    primary_result_keys,
    expected_detail,
):
    site = make_site("Alpha Campus")
    selected_result = make_result("site-1", "Alpha Campus", score=90.0)

    monkeypatch.setattr(routes_export, "get_site", lambda site_id: (site_id, site))

    config = routes_export.ReportConfig(
        report_type="executive",
        studied_site_ids=["site-1"],
        primary_result_keys=primary_result_keys,
        scenario_results=[selected_result.model_dump(mode="json")],
        layout_mode="presentation_16_9",
    )

    with pytest.raises(routes_export.HTTPException) as excinfo:
        asyncio.run(routes_export.export_html_endpoint(config))

    assert excinfo.value.status_code == 400
    assert expected_detail in excinfo.value.detail
